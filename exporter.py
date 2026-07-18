# for excel 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3
from datetime import datetime
#for pdf 
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

DB_PATH = "expenses.db"

# for excel 

def get_filtered_expenses(month=None, year=None, category=None):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    
    query = "SELECT * FROM expenses WHERE 1=1"
    params = []
    
    if month:
        query += " AND strftime('%m', date) = ?"
        params.append(str(month).zfill(2))
    
    if year:
        query += " AND strftime('%Y', date) = ?"
        params.append(str(year))
    
    if category:
        query += " AND category = ?"
        params.append(category)
    
    cursor = conn.execute(query, params)
    results = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return results

def generate_excel(expenses, filename="kharcha_export.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kharcha"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    
    headers = ["#", "Date", "Category", "Amount (Rs.)", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    total = 0
    for row_num, expense in enumerate(expenses, 2):
        ws.cell(row=row_num, column=1, value=row_num-1)
        ws.cell(row=row_num, column=2, value=expense.get("date"))
        ws.cell(row=row_num, column=3, value=expense.get("category"))
        amount = expense.get("amount", 0)
        ws.cell(row=row_num, column=4, value=amount)
        ws.cell(row=row_num, column=5, value=expense.get("description"))
        total += amount
    
    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30
    
    wb.save(filename)
    return filename
    
# for pdf 
def generate_pdf(expenses, filename="kharcha_export.pdf"):
    doc = SimpleDocTemplate(filename, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph("Kharcha — Expense Report", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [["#", "Date", "Category", "Amount (Rs.)", "Description"]]
    total = 0
    
    for i, expense in enumerate(expenses, 1):
        amount = expense.get("amount", 0)
        data.append([
            str(i),
            expense.get("date", ""),
            expense.get("category", ""),
            f"Rs. {amount:,}",
            expense.get("description", "")
        ])
        total += amount
    
    data.append(["", "", "TOTAL", f"Rs. {total:,}", ""])
    
    # Style
    table = Table(data, colWidths=[30, 80, 90, 90, 180])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2D6A4F')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F0F0F0')]),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (3,0), (3,-1), 'RIGHT'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filename